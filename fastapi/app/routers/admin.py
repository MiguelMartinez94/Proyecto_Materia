from datetime import datetime, timedelta, time
from typing import List, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case, extract
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models import (Usuario, ConfiguracionNegocio, Role, Ingrediente, Producto,
                        CategoriaMenu, Venta, Orden, DetalleOrden, Mesa, Gasto)
from ..schemas import (UsuarioOut, UsuarioCreate, ConfiguracionNegocioOut, ConfiguracionNegocioUpdate, BaseSchema,
                        IngredienteOut, ProductoOut, CategoriaMenuOut, IngredienteCreate, IngredienteUpdate,
                        ProductoCreate, ProductoUpdate, CategoriaMenuCreate, RoleOut)
from .auth import require_roles, get_password_hash

class UsuarioUpdate(BaseSchema):
    nombre: str | None = None
    username: str | None = None
    password: str | None = None
    rol_id: int | None = None
    activo: bool | None = None

router = APIRouter(
    prefix="/admin",
    tags=["Administración"],
    dependencies=[Depends(require_roles(["ADMINISTRADOR"]))]
)



@router.get("/usuarios", response_model=List[UsuarioOut])
def listar_usuarios(db: Session = Depends(get_db)):
    return db.query(Usuario).options(joinedload(Usuario.rol)).all()

@router.post("/usuarios", response_model=UsuarioOut)
def crear_usuario(user: UsuarioCreate, db: Session = Depends(get_db)):
    if db.query(Usuario).filter(Usuario.username == user.username).first():
        raise HTTPException(status_code=400, detail="El nombre de usuario ya existe")
    if not db.query(Role).filter(Role.id == user.rol_id).first():
        raise HTTPException(status_code=400, detail="Rol inválido")

    hashed_pw = get_password_hash(user.password)
    nuevo_usuario = Usuario(
        nombre=user.nombre,
        username=user.username,
        password_hash=hashed_pw,
        rol_id=user.rol_id,
        activo=user.activo
    )
    db.add(nuevo_usuario)
    db.commit()
    db.refresh(nuevo_usuario)
    return nuevo_usuario

@router.put("/usuarios/{id}", response_model=UsuarioOut)
def actualizar_usuario(id: int, user_update: UsuarioUpdate, db: Session = Depends(get_db)):
    usuario = db.query(Usuario).filter(Usuario.id == id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if user_update.username and user_update.username != usuario.username:
        if db.query(Usuario).filter(Usuario.username == user_update.username).first():
            raise HTTPException(status_code=400, detail="El nombre de usuario ya existe")
        usuario.username = user_update.username

    if user_update.nombre is not None:
        usuario.nombre = user_update.nombre
    if user_update.rol_id is not None:
        if not db.query(Role).filter(Role.id == user_update.rol_id).first():
            raise HTTPException(status_code=400, detail="Rol inválido")
        usuario.rol_id = user_update.rol_id
    if user_update.activo is not None:
        usuario.activo = user_update.activo
    if user_update.password is not None:
        usuario.password_hash = get_password_hash(user_update.password)

    db.commit()
    db.refresh(usuario)
    return usuario

@router.delete("/usuarios/{id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_usuario(id: int, db: Session = Depends(get_db)):
    usuario = db.query(Usuario).filter(Usuario.id == id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    try:
        db.delete(usuario)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="No se puede eliminar el usuario porque tiene registros asociados.")
    return None


@router.get("/configuracion", response_model=ConfiguracionNegocioOut)
def obtener_configuracion(db: Session = Depends(get_db)):
    config = db.query(ConfiguracionNegocio).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
    return config

@router.put("/configuracion", response_model=ConfiguracionNegocioOut)
def actualizar_configuracion(payload: ConfiguracionNegocioUpdate, db: Session = Depends(get_db)):
    config = db.query(ConfiguracionNegocio).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuración no encontrada")

    if payload.nombre_negocio is not None:
        config.nombre_negocio = payload.nombre_negocio
    if payload.mensaje_ticket is not None:
        config.mensaje_ticket = payload.mensaje_ticket
    if payload.impuesto_porcentaje is not None:
        config.impuesto_porcentaje = payload.impuesto_porcentaje
    if payload.moneda is not None:
        config.moneda = payload.moneda

    db.commit()
    db.refresh(config)
    return config


@router.get("/roles", response_model=List[RoleOut])
def listar_roles(db: Session = Depends(get_db)):
    return db.query(Role).all()

@router.get("/ingredientes", response_model=List[IngredienteOut])
def listar_ingredientes(db: Session = Depends(get_db)):
    return db.query(Ingrediente).order_by(Ingrediente.id).all()

@router.post("/ingredientes", response_model=IngredienteOut)
def crear_ingrediente(ingrediente: IngredienteCreate, db: Session = Depends(get_db)):
    nuevo = Ingrediente(**ingrediente.dict())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo

@router.put("/ingredientes/{id}", response_model=IngredienteOut)
def actualizar_ingrediente(id: int, payload: IngredienteUpdate, db: Session = Depends(get_db)):
    ingrediente = db.query(Ingrediente).filter(Ingrediente.id == id).first()
    if not ingrediente:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")

    update_data = payload.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(ingrediente, key, value)

    db.commit()
    db.refresh(ingrediente)
    return ingrediente

@router.delete("/ingredientes/{id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_ingrediente(id: int, db: Session = Depends(get_db)):
    ingrediente = db.query(Ingrediente).filter(Ingrediente.id == id).first()
    if not ingrediente:
        raise HTTPException(status_code=404, detail="Ingrediente no encontrado")
    db.delete(ingrediente)
    db.commit()
    return None

@router.get("/menu/categorias", response_model=List[CategoriaMenuOut])
def listar_categorias(db: Session = Depends(get_db)):
    return db.query(CategoriaMenu).all()

@router.post("/menu/categorias", response_model=CategoriaMenuOut)
def crear_categoria(categoria: CategoriaMenuCreate, db: Session = Depends(get_db)):
    nueva = CategoriaMenu(**categoria.dict())
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    return nueva

@router.get("/menu/productos", response_model=List[ProductoOut])
def listar_productos(db: Session = Depends(get_db)):
    return db.query(Producto).filter(Producto.disponible == True).all()

@router.post("/menu/productos", response_model=ProductoOut)
def crear_producto(producto: ProductoCreate, db: Session = Depends(get_db)):
    nuevo = Producto(**producto.dict())
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    return nuevo

@router.put("/menu/productos/{id}", response_model=ProductoOut)
def actualizar_producto(id: int, payload: ProductoUpdate, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    update_data = payload.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(producto, key, value)

    db.commit()
    db.refresh(producto)
    return producto

@router.delete("/menu/productos/{id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_producto(id: int, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    producto.disponible = False
    db.commit()
    return None


@router.get("/reportes/ventas")
def reporte_ventas(db: Session = Depends(get_db)):
    ventas = db.query(Venta).all()
    total = sum(v.total_pagado for v in ventas)
    return {
        "total_historico_vendido": round(total, 2),
        "total_ordenes_cobradas": len(ventas)
    }


@router.get("/reportes/dashboard")
def reporte_dashboard(db: Session = Depends(get_db)):
    hoy = datetime.utcnow().date()
    inicio_dia = datetime.combine(hoy, time.min)
    fin_dia = datetime.combine(hoy, time.max)

    ventas_hoy = db.query(Venta).filter(
        Venta.created_at >= inicio_dia,
        Venta.created_at <= fin_dia
    ).all()
    total_ventas_dia = sum(v.total_pagado for v in ventas_hoy)

    pedidos_pendientes = db.query(Orden).filter(Orden.estado != "PAGADA").count()

    ingredientes_criticos = db.query(Ingrediente).all()

    hace_7_dias = hoy - timedelta(days=6)
    inicio_semana = datetime.combine(hace_7_dias, time.min)

    ventas_semana = db.query(Venta).filter(
        Venta.created_at >= inicio_semana
    ).all()

    total_semanal = sum(v.total_pagado for v in ventas_semana)

    ventas_por_dia = {}
    dias_labels = []
    dias_nombres = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    for i in range(7):
        fecha = hace_7_dias + timedelta(days=i)
        dia_nombre = dias_nombres[fecha.weekday()]
        dias_labels.append(dia_nombre)
        ventas_por_dia[fecha.isoformat()] = 0.0

    for v in ventas_semana:
        fecha_key = v.created_at.date().isoformat()
        if fecha_key in ventas_por_dia:
            ventas_por_dia[fecha_key] += v.total_pagado

    ventas_diarias = [round(v, 2) for v in ventas_por_dia.values()]

    return {
        "ventas_dia": round(total_ventas_dia, 2),
        "pedidos_pendientes": pedidos_pendientes,
        "stock_critico_count": len(ingredientes_criticos),
        "ingredientes_criticos": [
            {
                "id": i.id,
                "nombre": i.nombre,
                "stock_actual": i.stock_actual,
                "stock_minimo": i.stock_minimo,
                "unidad_medida": i.unidad_medida
            } for i in ingredientes_criticos
        ],
        "ventas_semanales": round(total_semanal, 2),
        "grafica_dias_labels": dias_labels,
        "grafica_ventas_diarias": ventas_diarias
    }


@router.get("/reportes/estadisticas")
def reporte_estadisticas(
    periodo: str = Query("semana", regex="^(hoy|semana|mes)$"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    db: Session = Depends(get_db)
):
    hoy = datetime.utcnow().date()

    if fecha_inicio and fecha_fin:
        try:
            f_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            f_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    elif periodo == "hoy":
        f_inicio = hoy
        f_fin = hoy
    elif periodo == "semana":
        f_inicio = hoy - timedelta(days=6)
        f_fin = hoy
    else:  # mes
        f_inicio = hoy.replace(day=1)
        f_fin = hoy

    dt_inicio = datetime.combine(f_inicio, time.min)
    dt_fin = datetime.combine(f_fin, time.max)

    ventas = db.query(Venta).filter(
        Venta.created_at >= dt_inicio,
        Venta.created_at <= dt_fin
    ).all()

    total_ventas = sum(v.total_pagado for v in ventas)
    num_ventas = len(ventas)
    ticket_promedio = (total_ventas / num_ventas) if num_ventas > 0 else 0.0

    dias_rango = (f_fin - f_inicio).days + 1
    ventas_por_dia_labels = []
    ventas_por_dia_data = []
    dias_nombres = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    fechas_map = {}
    for i in range(dias_rango):
        fecha = f_inicio + timedelta(days=i)
        label = dias_nombres[fecha.weekday()] if dias_rango <= 7 else fecha.strftime("%d/%m")
        ventas_por_dia_labels.append(label)
        fechas_map[fecha.isoformat()] = i
        ventas_por_dia_data.append(0.0)

    for v in ventas:
        key = v.created_at.date().isoformat()
        if key in fechas_map:
            ventas_por_dia_data[fechas_map[key]] += v.total_pagado

    ventas_por_dia_data = [round(x, 2) for x in ventas_por_dia_data]

    productos_vendidos = db.query(
        Producto.nombre,
        CategoriaMenu.nombre.label("categoria"),
        func.sum(DetalleOrden.cantidad).label("total_vendido")
    ).join(
        DetalleOrden, DetalleOrden.producto_id == Producto.id
    ).join(
        Orden, Orden.id == DetalleOrden.orden_id
    ).outerjoin(
        CategoriaMenu, CategoriaMenu.id == Producto.categoria_id
    ).filter(
        Orden.created_at >= dt_inicio,
        Orden.created_at <= dt_fin,
        Orden.estado == "PAGADA"
    ).group_by(Producto.nombre, CategoriaMenu.nombre).order_by(
        func.sum(DetalleOrden.cantidad).desc()
    ).limit(5).all()

    total_items_vendidos = sum(p.total_vendido for p in productos_vendidos) if productos_vendidos else 1
    productos_top = []
    for p in productos_vendidos:
        porcentaje = round((p.total_vendido / total_items_vendidos) * 100, 1) if total_items_vendidos > 0 else 0
        productos_top.append({
            "nombre": p.nombre,
            "categoria": p.categoria or "General",
            "cantidad": p.total_vendido,
            "porcentaje": porcentaje
        })

    ordenes = db.query(Orden).options(
        joinedload(Orden.mesa),
        joinedload(Orden.mesero),
        joinedload(Orden.items).joinedload(DetalleOrden.producto)
    ).filter(
        Orden.created_at >= dt_inicio,
        Orden.created_at <= dt_fin
    ).order_by(Orden.created_at.desc()).all()

    pedidos_detalle = []
    for o in ordenes:
        pedidos_detalle.append({
            "id": o.id,
            "mesa_numero": o.mesa.numero if o.mesa else None,
            "mesero": o.mesero.nombre if o.mesero else "N/A",
            "estado": o.estado,
            "total": round(o.total, 2),
            "num_items": sum(item.cantidad for item in o.items),
            "created_at": o.created_at.isoformat(),
            "items": [
                {
                    "producto": item.producto.nombre if item.producto else "N/A",
                    "cantidad": item.cantidad,
                    "precio_unitario": item.precio_unitario,
                    "subtotal": round(item.cantidad * item.precio_unitario, 2)
                } for item in o.items
            ]
        })

    gastos_db = db.query(Gasto).filter(
        Gasto.created_at >= dt_inicio,
        Gasto.created_at <= dt_fin
    ).order_by(Gasto.created_at.desc()).all()

    gastos_detalle = []
    for g in gastos_db:
        gastos_detalle.append({
            "descripcion": g.descripcion,
            "monto": round(g.monto, 2),
            "created_at": g.created_at.isoformat()
        })

    return {
        "periodo": {"inicio": f_inicio.isoformat(), "fin": f_fin.isoformat()},
        "resumen": {
            "total_ventas": round(total_ventas, 2),
            "num_ventas": num_ventas,
            "ticket_promedio": round(ticket_promedio, 2)
        },
        "grafica_ventas": {
            "labels": ventas_por_dia_labels,
            "data": ventas_por_dia_data
        },
        "productos_top": productos_top,
        "pedidos": pedidos_detalle,
        "gastos": gastos_detalle
    }



@router.get("/reportes/export/pdf")
def exportar_pdf(
    periodo: str = Query("semana", regex="^(hoy|semana|mes)$"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    tipo: str = Query("ventas", regex="^(ventas|pedidos|inventario|gastos)$"),
    db: Session = Depends(get_db)
):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    hoy = datetime.utcnow().date()

    if fecha_inicio and fecha_fin:
        try:
            f_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            f_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato inválido")
    elif periodo == "hoy":
        f_inicio = f_fin = hoy
    elif periodo == "semana":
        f_inicio = hoy - timedelta(days=6)
        f_fin = hoy
    else:
        f_inicio = hoy.replace(day=1)
        f_fin = hoy

    dt_inicio = datetime.combine(f_inicio, time.min)
    dt_fin = datetime.combine(f_fin, time.max)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=12)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.grey, spaceAfter=20)
    elements = []

    header_color = colors.HexColor("#4a3b2c")
    alt_row = colors.HexColor("#f5ece4")

    if tipo == "ventas":
        elements.append(Paragraph("Reporte de Ventas", title_style))
        elements.append(Paragraph(f"Periodo: {f_inicio.strftime('%d/%m/%Y')} - {f_fin.strftime('%d/%m/%Y')}", subtitle_style))

        ventas = db.query(Venta).options(
            joinedload(Venta.orden).joinedload(Orden.mesa)
        ).filter(
            Venta.created_at >= dt_inicio, Venta.created_at <= dt_fin
        ).order_by(Venta.created_at.desc()).all()

        data = [["Folio", "Fecha", "Mesa", "Método Pago", "Total", "Recibido", "Cambio"]]
        total_general = 0
        for v in ventas:
            mesa_num = v.orden.mesa.numero if v.orden and v.orden.mesa else "N/A"
            data.append([
                v.ticket_folio,
                v.created_at.strftime("%d/%m/%Y %H:%M"),
                str(mesa_num),
                v.metodo_pago,
                f"${v.total_pagado:,.2f}",
                f"${v.monto_recibido:,.2f}",
                f"${v.cambio:,.2f}"
            ])
            total_general += v.total_pagado

        data.append(["", "", "", "TOTAL:", f"${total_general:,.2f}", "", ""])

    elif tipo == "pedidos":
        elements.append(Paragraph("Reporte de Pedidos", title_style))
        elements.append(Paragraph(f"Periodo: {f_inicio.strftime('%d/%m/%Y')} - {f_fin.strftime('%d/%m/%Y')}", subtitle_style))

        ordenes = db.query(Orden).options(
            joinedload(Orden.mesa), joinedload(Orden.mesero),
            joinedload(Orden.items).joinedload(DetalleOrden.producto)
        ).filter(
            Orden.created_at >= dt_inicio, Orden.created_at <= dt_fin
        ).order_by(Orden.created_at.desc()).all()

        data = [["# Orden", "Fecha", "Mesa", "Mesero", "Estado", "Items", "Total"]]
        for o in ordenes:
            data.append([
                str(o.id),
                o.created_at.strftime("%d/%m/%Y %H:%M"),
                str(o.mesa.numero) if o.mesa else "N/A",
                o.mesero.nombre if o.mesero else "N/A",
                o.estado,
                str(sum(i.cantidad for i in o.items)),
                f"${o.total:,.2f}"
            ])

    elif tipo == "gastos":
        elements.append(Paragraph("Reporte de Gastos", title_style))
        elements.append(Paragraph(f"Periodo: {f_inicio.strftime('%d/%m/%Y')} - {f_fin.strftime('%d/%m/%Y')}", subtitle_style))

        gastos = db.query(Gasto).filter(
            Gasto.created_at >= dt_inicio, Gasto.created_at <= dt_fin
        ).order_by(Gasto.created_at.desc()).all()

        data = [["Fecha", "Descripción", "Monto"]]
        total_gastos = 0
        for g in gastos:
            data.append([
                g.created_at.strftime("%d/%m/%Y %H:%M"),
                g.descripcion,
                f"${g.monto:,.2f}"
            ])
            total_gastos += g.monto

        data.append(["", "TOTAL GASTOS:", f"${total_gastos:,.2f}"])

    else:  
        elements.append(Paragraph("Reporte de Inventario", title_style))
        elements.append(Paragraph(f"Generado: {hoy.strftime('%d/%m/%Y')}", subtitle_style))

        ingredientes = db.query(Ingrediente).order_by(Ingrediente.nombre).all()
        data = [["Insumo", "Unidad", "Stock Actual", "Stock Mínimo", "Estado"]]
        for i in ingredientes:
            estado = "CRÍTICO" if i.stock_actual <= i.stock_minimo else "OK"
            unidad = "pzas" if i.unidad_medida == "pcs" else i.unidad_medida
            data.append([
                i.nombre,
                unidad,
                str(i.stock_actual),
                str(i.stock_minimo),
                estado
            ])

    col_count = len(data[0])
    table = Table(data, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e3d3c4")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_row]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"reporte_{tipo}_{f_inicio}_{f_fin}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@router.get("/reportes/export/xlsx")
def exportar_xlsx(
    periodo: str = Query("semana", regex="^(hoy|semana|mes)$"),
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    tipo: str = Query("ventas", regex="^(ventas|pedidos|inventario|gastos)$"),
    db: Session = Depends(get_db)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    hoy = datetime.utcnow().date()

    if fecha_inicio and fecha_fin:
        try:
            f_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            f_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato inválido")
    elif periodo == "hoy":
        f_inicio = f_fin = hoy
    elif periodo == "semana":
        f_inicio = hoy - timedelta(days=6)
        f_fin = hoy
    else:
        f_inicio = hoy.replace(day=1)
        f_fin = hoy

    dt_inicio = datetime.combine(f_inicio, time.min)
    dt_fin = datetime.combine(f_fin, time.max)

    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4A3B2C", end_color="4A3B2C", fill_type="solid")
    alt_fill = PatternFill(start_color="F5ECE4", end_color="F5ECE4", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='E3D3C4'),
        right=Side(style='thin', color='E3D3C4'),
        top=Side(style='thin', color='E3D3C4'),
        bottom=Side(style='thin', color='E3D3C4')
    )

    if tipo == "ventas":
        ws.title = "Ventas"
        headers = ["Folio", "Fecha", "Mesa", "Método Pago", "Total", "Recibido", "Cambio"]

        ventas = db.query(Venta).options(
            joinedload(Venta.orden).joinedload(Orden.mesa)
        ).filter(
            Venta.created_at >= dt_inicio, Venta.created_at <= dt_fin
        ).order_by(Venta.created_at.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, v in enumerate(ventas, 2):
            mesa_num = v.orden.mesa.numero if v.orden and v.orden.mesa else "N/A"
            values = [
                v.ticket_folio,
                v.created_at.strftime("%d/%m/%Y %H:%M"),
                mesa_num,
                v.metodo_pago,
                v.total_pagado,
                v.monto_recibido,
                v.cambio
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    elif tipo == "pedidos":
        ws.title = "Pedidos"
        headers = ["# Orden", "Fecha", "Mesa", "Mesero", "Estado", "Items", "Total"]

        ordenes = db.query(Orden).options(
            joinedload(Orden.mesa), joinedload(Orden.mesero),
            joinedload(Orden.items).joinedload(DetalleOrden.producto)
        ).filter(
            Orden.created_at >= dt_inicio, Orden.created_at <= dt_fin
        ).order_by(Orden.created_at.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, o in enumerate(ordenes, 2):
            values = [
                o.id,
                o.created_at.strftime("%d/%m/%Y %H:%M"),
                o.mesa.numero if o.mesa else "N/A",
                o.mesero.nombre if o.mesero else "N/A",
                o.estado,
                sum(i.cantidad for i in o.items),
                o.total
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    elif tipo == "gastos":
        ws.title = "Gastos"
        headers = ["Fecha", "Descripción", "Monto"]

        gastos = db.query(Gasto).filter(
            Gasto.created_at >= dt_inicio, Gasto.created_at <= dt_fin
        ).order_by(Gasto.created_at.desc()).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, g in enumerate(gastos, 2):
            values = [
                g.created_at.strftime("%d/%m/%Y %H:%M"),
                g.descripcion,
                g.monto
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    else:  
        ws.title = "Inventario"
        headers = ["Insumo", "Unidad", "Stock Actual", "Stock Mínimo", "Estado"]

        ingredientes = db.query(Ingrediente).order_by(Ingrediente.nombre).all()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, i in enumerate(ingredientes, 2):
            estado = "CRÍTICO" if i.stock_actual <= i.stock_minimo else "OK"
            unidad = "pzas" if i.unidad_medida == "pcs" else i.unidad_medida
            values = [i.nombre, unidad, i.stock_actual, i.stock_minimo, estado]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                if col == 5 and val == "CRÍTICO":
                    cell.font = Font(bold=True, color="D32F2F")

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 4
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_{tipo}_{f_inicio}_{f_fin}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
